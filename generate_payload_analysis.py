"""
Generate an Excel analysis of the PAFS submission payload.

Columns:
  A  Section
  B  Property Name          (key in JSON payload)
  C  Full Payload Path      (dot-notation)
  D  Data Type              (string, integer, decimal, boolean, array, object)
  E  Schema Required?       (Required / Optional)
  F  Conditional?           (Yes / No)
  G  Condition Description  (when/why the field value changes)
  H  Value When Condition NOT Met (or "Field omitted" / "n/a")
  I  Source Field(s) in Project Object
  J  Notes / Allowed Values
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Data
# ---------------------------------------------------------------------------

COLUMNS = [
    "Section",
    "Property Name",
    "Full Payload Path",
    "Data Type",
    "Schema Required?",
    "Conditional?",
    "Condition Description",
    "Value When Condition Not Met",
    "Source Field(s) in Project Object",
    "Notes / Allowed Values",
]

# Each row: (section, property_name, path, data_type, schema_required,
#            conditional, condition_desc, value_when_not_met, source_fields, notes)
ROWS = [
    # ── Core Identity ────────────────────────────────────────────────────────
    (
        "Core Identity",
        "name",
        "name",
        "string",
        "Required",
        "No",
        "—",
        "n/a",
        "project.name",
        "Human-readable project title",
    ),
    (
        "Core Identity",
        "type",
        "type",
        "string",
        "Required",
        "No",
        "—",
        "n/a",
        "project.projectType",
        "DEF | REP | REF | HCR | STR | STU | ELO",
    ),
    (
        "Core Identity",
        "main_intervention_type",
        "main_intervention_type",
        "string | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.mainInterventionType",
        "Label-mapped via MAIN_INTERVENTION_TYPE_LABELS. "
        "Only populated for DEF/REF/REP types.",
    ),
    (
        "Core Identity",
        "natural_flood_management",
        "intervention_types.natural_flood_management",
        "boolean",
        "Optional (sub-required)",
        "No",
        "—",
        "false",
        "project.projectInterventionTypes (CSV)",
        "true when 'nfm' present in the comma-separated list",
    ),
    (
        "Core Identity",
        "property_flood_resilience",
        "intervention_types.property_flood_resilience",
        "boolean",
        "Optional (sub-required)",
        "No",
        "—",
        "false",
        "project.projectInterventionTypes (CSV)",
        "true when 'pfr' present",
    ),
    (
        "Core Identity",
        "sustainable_drainage_systems",
        "intervention_types.sustainable_drainage_systems",
        "boolean",
        "Optional (sub-required)",
        "No",
        "—",
        "false",
        "project.projectInterventionTypes (CSV)",
        "true when 'suds' present",
    ),
    (
        "Core Identity",
        "other",
        "intervention_types.other",
        "boolean",
        "Optional (sub-required)",
        "No",
        "—",
        "false",
        "project.projectInterventionTypes (CSV)",
        "true when 'other' present",
    ),
    (
        "Core Identity",
        "national_project_number",
        "national_project_number",
        "string | null",
        "Required",
        "No",
        "—",
        "null",
        "project.referenceNumber",
        "e.g. AC501E/001A/001A",
    ),
    (
        "Core Identity",
        "pafs_region_and_coastal_commitee",
        "pafs_region_and_coastal_commitee",
        "string | null",
        "Required",
        "Yes",
        "referenceNumber must be present and have a known 2-char RFCC prefix",
        "null",
        "Derived from first 2 chars of project.referenceNumber "
        "via RFCC_CODE_NAMES map",
        "e.g. 'AC' → 'Anglian (Great Ouse)'. null if prefix unrecognised.",
    ),
    (
        "Core Identity",
        "pafs_ea_area",
        "pafs_ea_area",
        "string | null",
        "Required",
        "No",
        "—",
        "null",
        "project.eaAreaName",
        "",
    ),
    (
        "Core Identity",
        "lrma_name",
        "lrma_name",
        "string | null",
        "Required",
        "No",
        "—",
        "null",
        "project.rmaName",
        "",
    ),
    (
        "Core Identity",
        "lrma_type",
        "lrma_type",
        "string | null",
        "Required",
        "No",
        "—",
        "null",
        "project.rmaSubType",
        "",
    ),
    (
        "Core Identity",
        "email",
        "email",
        "string | null",
        "Optional",
        "No",
        "—",
        "null",
        "creatorEmail parameter (looked up from pafs_core_users via creator_id)",
        "Looked up just before submission; null if user not found",
    ),
    (
        "Core Identity",
        "shapefile",
        "shapefile",
        "string | null",
        "Required",
        "Yes",
        "Benefit area file must exist in S3 (project.benefitAreaFileName)",
        "null",
        "shapefileBase64 param — fetched from S3 at submit time",
        "Base-64 encoded shapefile. null if S3 fetch fails or no file.",
    ),

    # ── Gateway Dates ────────────────────────────────────────────────────────
    (
        "Gateway Dates",
        "aspirational_gateway_1",
        "aspirational_gateway_1",
        "MM/YYYY | null",
        "Required",
        "Yes",
        "Both startOutlineBusinessCaseMonth AND startOutlineBusinessCaseYear "
        "must be present",
        "null",
        "project.startOutlineBusinessCaseMonth, project.startOutlineBusinessCaseYear",
        "Gateway 1 — Start OBC. Formatted as zero-padded MM/YYYY.",
    ),
    (
        "Gateway Dates",
        "aspirational_gateway_2",
        "aspirational_gateway_2",
        "MM/YYYY | null",
        "Required",
        "Yes",
        "Both completeOutlineBusinessCaseMonth AND completeOutlineBusinessCaseYear "
        "must be present",
        "null",
        "project.completeOutlineBusinessCaseMonth, project.completeOutlineBusinessCaseYear",
        "Gateway 2 — Complete OBC.",
    ),
    (
        "Gateway Dates",
        "aspirational_gateway_3",
        "aspirational_gateway_3",
        "MM/YYYY | null",
        "Required",
        "Yes",
        "Both awardContractMonth AND awardContractYear must be present",
        "null",
        "project.awardContractMonth, project.awardContractYear",
        "Gateway 3 — Award Contract.",
    ),
    (
        "Gateway Dates",
        "aspirational_gateway_4",
        "aspirational_gateway_4",
        "MM/YYYY | null",
        "Required",
        "Yes",
        "Both readyForServiceMonth AND readyForServiceYear must be present",
        "null",
        "project.readyForServiceMonth, project.readyForServiceYear",
        "Gateway 4 — Ready for Service.",
    ),
    (
        "Gateway Dates",
        "aspirational_start_of_construction",
        "aspirational_start_of_construction",
        "MM/YYYY | null",
        "Required",
        "Yes",
        "Both startConstructionMonth AND startConstructionYear must be present",
        "null",
        "project.startConstructionMonth, project.startConstructionYear",
        "Start of Construction date.",
    ),
    (
        "Gateway Dates",
        "earliest_start_date_with_gia_available",
        "earliest_start_date_with_gia_available",
        "MM/YYYY | null",
        "Required",
        "Yes",
        "couldStartEarly === true AND both earliestWithGiaMonth AND "
        "earliestWithGiaYear must be present",
        "null",
        "project.earliestWithGiaMonth, project.earliestWithGiaYear",
        "Only meaningful when project could start earlier with GIA funding.",
    ),
    (
        "Gateway Dates",
        "earliest_start_date",
        "earliest_start_date",
        "MM/YYYY | null",
        "Required",
        "Yes",
        "project.financialStartYear must be present",
        "null",
        "project.financialStartYear",
        "Always formatted as '04/<financialStartYear>' (April = financial year start).",
    ),

    # ── Secondary Risk Sources ───────────────────────────────────────────────
    (
        "Secondary Risk Sources",
        "fluvial_flooding",
        "secondary_risk_sources.fluvial_flooding",
        "boolean",
        "Required",
        "No",
        "—",
        "false",
        "project.risks or project.projectRisksProtectedAgainst (CSV parsed)",
        "true when 'fluvial_flooding' is in the comma-separated risks string",
    ),
    (
        "Secondary Risk Sources",
        "tidal_flooding",
        "secondary_risk_sources.tidal_flooding",
        "boolean",
        "Required",
        "No",
        "—",
        "false",
        "project.risks or project.projectRisksProtectedAgainst (CSV)",
        "",
    ),
    (
        "Secondary Risk Sources",
        "groundwater_flooding",
        "secondary_risk_sources.groundwater_flooding",
        "boolean",
        "Required",
        "No",
        "—",
        "false",
        "project.risks or project.projectRisksProtectedAgainst (CSV)",
        "",
    ),
    (
        "Secondary Risk Sources",
        "surface_water_flooding",
        "secondary_risk_sources.surface_water_flooding",
        "boolean",
        "Required",
        "No",
        "—",
        "false",
        "project.risks or project.projectRisksProtectedAgainst (CSV)",
        "",
    ),
    (
        "Secondary Risk Sources",
        "sea_flooding",
        "secondary_risk_sources.sea_flooding",
        "boolean",
        "Required",
        "No",
        "—",
        "false",
        "project.risks or project.projectRisksProtectedAgainst (CSV)",
        "",
    ),
    (
        "Secondary Risk Sources",
        "reservoir_flooding",
        "secondary_risk_sources.reservoir_flooding",
        "boolean",
        "Required",
        "No",
        "—",
        "false",
        "project.risks or project.projectRisksProtectedAgainst (CSV)",
        "",
    ),
    (
        "Secondary Risk Sources",
        "coastal_erosion",
        "secondary_risk_sources.coastal_erosion",
        "boolean",
        "Required",
        "No",
        "—",
        "false",
        "project.risks or project.projectRisksProtectedAgainst (CSV)",
        "",
    ),

    # ── Risk & Properties ────────────────────────────────────────────────────
    (
        "Risk & Properties",
        "risk_source",
        "risk_source",
        "string | null",
        "Required",
        "No",
        "—",
        "null",
        "project.mainRisk",
        "Label-mapped via RISK_LABELS (fcerm1-labels). null if mainRisk is null.",
    ),
    (
        "Risk & Properties",
        "properties_benefitting_in_20pct_most_deprived_areas",
        "properties_benefitting_in_20pct_most_deprived_areas",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.percentProperties20PercentDeprived",
        "0–100 percentage. Required by submission validation.",
    ),
    (
        "Risk & Properties",
        "properties_benefitting_in_40pct_most_deprived_areas",
        "properties_benefitting_in_40pct_most_deprived_areas",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.percentProperties40PercentDeprived",
        "0–100 percentage. Required by submission validation.",
    ),
    (
        "Risk & Properties",
        "fluvial_and_tidal_flood_risk",
        "fluvial_and_tidal_flood_risk",
        "string | null",
        "Optional",
        "Yes",
        "currentFloodFluvialRisk is null → defaults to 'not_applicable' "
        "before label lookup",
        "Label for 'not_applicable' (i.e. the raw value 'not_applicable', "
        "as it has no entry in the map)",
        "project.currentFloodFluvialRisk",
        "PAYLOAD_FLOOD_RISK_LABELS: high | medium | low | very_low. "
        "Required by validation only when fluvial/tidal/sea risk selected.",
    ),
    (
        "Risk & Properties",
        "surface_water_flood_risk",
        "surface_water_flood_risk",
        "string | null",
        "Optional",
        "Yes",
        "currentFloodSurfaceWaterRisk is null → defaults to 'not_applicable'",
        "'not_applicable'",
        "project.currentFloodSurfaceWaterRisk",
        "Required by validation only when surface_water risk selected.",
    ),
    (
        "Risk & Properties",
        "coastal_erosion_flood_risk",
        "coastal_erosion_flood_risk",
        "string | null",
        "Optional",
        "Yes",
        "currentCoastalErosionRisk is null → defaults to 'not_applicable'",
        "'not_applicable'",
        "project.currentCoastalErosionRisk",
        "PAYLOAD_COASTAL_EROSION_LABELS: medium_term | longer_term. "
        "Required by validation only when coastal_erosion risk selected.",
    ),

    # ── Goals & Approach ─────────────────────────────────────────────────────
    (
        "Goals & Approach",
        "problem_and_proposed_solution",
        "problem_and_proposed_solution",
        "string | null",
        "Required",
        "No",
        "—",
        "null",
        "project.approach",
        "Free-text description of the problem and proposed solution.",
    ),
    (
        "Goals & Approach",
        "moderation_code",
        "moderation_code",
        "string | null",
        "Required",
        "No",
        "—",
        "null",
        "project.urgencyReason",
        "Label-mapped via MODERATION_LABELS (e.g. 'statutory_need' → 'BS'). "
        "null if urgencyReason is null.",
    ),
    (
        "Goals & Approach",
        "urgency_details",
        "urgency_details",
        "string | null",
        "Optional",
        "Yes",
        "urgencyReason is NOT 'not_urgent' (i.e. statutory_need, legal_need, "
        "health_and_safety, emergency_works, time_limited) → urgency_details required",
        "null",
        "project.urgencyDetails",
        "Submission validation blocks if urgencyReason is urgent "
        "but urgencyDetails is empty.",
    ),

    # ── Outcome Measures — OM2 ───────────────────────────────────────────────
    (
        "Outcome Measures — OM2",
        "om2.1",
        "outcome_measures.om2.om2.1",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.maintainingExistingAssets",
        "Properties at risk: maintaining existing assets (flood).",
    ),
    (
        "Outcome Measures — OM2",
        "om2.2",
        "outcome_measures.om2.om2.2",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.reducingFloodRisk50Plus",
        "Properties benefitting: reducing flood risk by ≥50%.",
    ),
    (
        "Outcome Measures — OM2",
        "om2.3",
        "outcome_measures.om2.om2.3",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.reducingFloodRiskLess50",
        "Properties benefitting: reducing flood risk by <50%.",
    ),
    (
        "Outcome Measures — OM2",
        "om2.4",
        "outcome_measures.om2.om2.4",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.increasingFloodResilience",
        "Properties benefitting: increasing flood resilience.",
    ),

    # ── Outcome Measures — OM3 ───────────────────────────────────────────────
    (
        "Outcome Measures — OM3",
        "om3.1",
        "outcome_measures.om3.om3.1",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.propertiesBenefitMaintainingAssetsCoastal",
        "Coastal: maintaining existing assets.",
    ),
    (
        "Outcome Measures — OM3",
        "om3.2",
        "outcome_measures.om3.om3.2",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.propertiesBenefitInvestmentCoastalErosion",
        "Coastal: investment in coastal erosion.",
    ),

    # ── Outcome Measures — OM4a (Habitat) ────────────────────────────────────
    (
        "Outcome Measures — OM4a",
        "om4a_hectares_intertidal",
        "outcome_measures.om4a.om4a_hectares_intertidal",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.hectaresOfIntertidalHabitatCreatedOrEnhanced",
        "Habitat area in hectares. Up to 16 digits + 2 d.p.",
    ),
    (
        "Outcome Measures — OM4a",
        "om4a_hectares_woodland",
        "outcome_measures.om4a.om4a_hectares_woodland",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.hectaresOfWoodlandHabitatCreatedOrEnhanced",
        "",
    ),
    (
        "Outcome Measures — OM4a",
        "om4a_hectares_wet_woodland",
        "outcome_measures.om4a.om4a_hectares_wet_woodland",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.hectaresOfWetWoodlandHabitatCreatedOrEnhanced",
        "",
    ),
    (
        "Outcome Measures — OM4a",
        "om4a_hectares_wetland_or_wet_grassland",
        "outcome_measures.om4a.om4a_hectares_wetland_or_wet_grassland",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.hectaresOfWetlandOrWetGrasslandCreatedOrEnhanced",
        "",
    ),
    (
        "Outcome Measures — OM4a",
        "om4a_hectares_grassland",
        "outcome_measures.om4a.om4a_hectares_grassland",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.hectaresOfGrasslandHabitatCreatedOrEnhanced",
        "",
    ),
    (
        "Outcome Measures — OM4a",
        "om4a_hectares_heathland",
        "outcome_measures.om4a.om4a_hectares_heathland",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.hectaresOfHeathlandCreatedOrEnhanced",
        "",
    ),
    (
        "Outcome Measures — OM4a",
        "om4a_hectares_ponds_lakes",
        "outcome_measures.om4a.om4a_hectares_ponds_lakes",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.hectaresOfPondOrLakeHabitatCreatedOrEnhanced",
        "",
    ),
    (
        "Outcome Measures — OM4a",
        "om4a_hectares_arable_land",
        "outcome_measures.om4a.om4a_hectares_arable_land",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.hectaresOfArableLandLakeHabitatCreatedOrEnhanced",
        "",
    ),

    # ── Outcome Measures — OM4b (Watercourse) ────────────────────────────────
    (
        "Outcome Measures — OM4b",
        "om4b_kilometres_of_watercourse_comprehensive",
        "outcome_measures.om4b.om4b_kilometres_of_watercourse_comprehensive",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.kilometresOfWatercourseEnhancedOrCreatedComprehensive",
        "Kilometres of watercourse (comprehensive management). Up to 16 digits + 2 d.p.",
    ),
    (
        "Outcome Measures — OM4b",
        "om4b_kilometres_of_watercourse_partial",
        "outcome_measures.om4b.om4b_kilometres_of_watercourse_partial",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.kilometresOfWatercourseEnhancedOrCreatedPartial",
        "Partial management.",
    ),
    (
        "Outcome Measures — OM4b",
        "om4b_kilometres_of_watercourse_single",
        "outcome_measures.om4b.om4b_kilometres_of_watercourse_single",
        "decimal | null",
        "Required (nullable)",
        "No",
        "—",
        "null",
        "project.kilometresOfWatercourseEnhancedOrCreatedSingle",
        "Single-sided management.",
    ),

    # ── Confidence ───────────────────────────────────────────────────────────
    (
        "Confidence",
        "homes_better_protected",
        "confidence.homes_better_protected",
        "string | null",
        "Required (nullable)",
        "Yes",
        "Submission validation requires a non-null value only for "
        "MANDATORY_WL_TYPES (DEF, REF, REP). For HCR, ELO, STR, STU it may "
        "be null and validation passes.",
        "null",
        "project.confidenceHomesBetterProtected",
        "PAYLOAD_CONFIDENCE_LABELS: high | medium_high | medium_low | low | "
        "not_applicable",
    ),
    (
        "Confidence",
        "homes_by_gateway_four",
        "confidence.homes_by_gateway_four",
        "string | null",
        "Required (nullable)",
        "Yes",
        "Non-null only required for MANDATORY_WL_TYPES (DEF, REF, REP)",
        "null",
        "project.confidenceHomesByGatewayFour",
        "Same allowed values as homes_better_protected.",
    ),
    (
        "Confidence",
        "secured_partnership_funding",
        "confidence.secured_partnership_funding",
        "string | null",
        "Required (nullable)",
        "Yes",
        "Non-null only required for MANDATORY_WL_TYPES (DEF, REF, REP)",
        "null",
        "project.confidenceSecuredPartnershipFunding",
        "Same allowed values as homes_better_protected.",
    ),

    # ── NFM Details ──────────────────────────────────────────────────────────
    (
        "NFM Details",
        "landowner_consent",
        "landowner_consent",
        "string | null",
        "Optional",
        "Yes",
        "Populated only when NFM or SUDS intervention types are selected; "
        "submission validation enforces the NFM section only for NFM/SUDS types",
        "null",
        "project.nfmLandownerConsent",
        "LANDOWNER_CONSENT_LABELS: Consent fully secured | Engaged but not "
        "fully secured | Initial contact made | Not yet engaged",
    ),
    (
        "NFM Details",
        "experience_of_nfm_measures",
        "experience_of_nfm_measures",
        "string | null",
        "Optional",
        "Yes",
        "NFM or SUDS intervention types selected",
        "null",
        "project.nfmExperienceLevel",
        "EXPERIENCE_LEVEL_LABELS: No experience | Some experience | Moderate "
        "experience | Extensive experience",
    ),
    (
        "NFM Details",
        "how_developed_is_the_proposal",
        "how_developed_is_the_proposal",
        "string | null",
        "Optional",
        "Yes",
        "NFM or SUDS intervention types selected",
        "null",
        "project.nfmProjectReadiness",
        "PROJECT_READINESS_LABELS: Early concept | Developing proposal | Well "
        "developed proposal | Ready to deliver",
    ),

    # ── NFM Measures (conditionally included) ────────────────────────────────
    (
        "NFM Measures",
        "river_and_floodplain_area",
        "river_and_floodplain_area",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'river_floodplain_restoration' must exist in "
        "pafs_core_nfm_measures array",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='river_floodplain_restoration'].areaHectares",
        "Area in hectares.",
    ),
    (
        "NFM Measures",
        "river_and_floodplain_volume",
        "river_and_floodplain_volume",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'river_floodplain_restoration' must exist in "
        "pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='river_floodplain_restoration'].storageVolumeM3",
        "Volume in m³.",
    ),
    (
        "NFM Measures",
        "leaky_barriers_volume",
        "leaky_barriers_volume",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'leaky_barriers' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='leaky_barriers'].storageVolumeM3",
        "",
    ),
    (
        "NFM Measures",
        "leaky_barriers_length",
        "leaky_barriers_length",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'leaky_barriers' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='leaky_barriers'].lengthKm",
        "",
    ),
    (
        "NFM Measures",
        "leaky_barriers_width",
        "leaky_barriers_width",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'leaky_barriers' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='leaky_barriers'].widthM",
        "",
    ),
    (
        "NFM Measures",
        "offline_storage_area",
        "offline_storage_area",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'offline_storage' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='offline_storage'].areaHectares",
        "",
    ),
    (
        "NFM Measures",
        "offline_storage_volume",
        "offline_storage_volume",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'offline_storage' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='offline_storage'].storageVolumeM3",
        "",
    ),
    (
        "NFM Measures",
        "woodland_area",
        "woodland_area",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'woodland' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='woodland'].areaHectares",
        "",
    ),
    (
        "NFM Measures",
        "headwater_area",
        "headwater_area",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'headwater_drainage' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='headwater_drainage'].areaHectares",
        "",
    ),
    (
        "NFM Measures",
        "runoff_attenuation_area",
        "runoff_attenuation_area",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'runoff_management' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='runoff_management'].areaHectares",
        "",
    ),
    (
        "NFM Measures",
        "runoff_attenuation_volume",
        "runoff_attenuation_volume",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'runoff_management' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='runoff_management'].storageVolumeM3",
        "",
    ),
    (
        "NFM Measures",
        "saltmarsh_area",
        "saltmarsh_area",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'saltmarsh_management' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='saltmarsh_management'].areaHectares",
        "",
    ),
    (
        "NFM Measures",
        "saltmarsh_length",
        "saltmarsh_length",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'saltmarsh_management' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='saltmarsh_management'].lengthKm",
        "",
    ),
    (
        "NFM Measures",
        "sand_dune_area",
        "sand_dune_area",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'sand_dune_management' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='sand_dune_management'].areaHectares",
        "",
    ),
    (
        "NFM Measures",
        "sand_dune_length",
        "sand_dune_length",
        "decimal | null",
        "Optional",
        "Yes",
        "Measure type 'sand_dune_management' must exist in pafs_core_nfm_measures",
        "Field OMITTED from payload",
        "nfmMeasures[measureType='sand_dune_management'].lengthKm",
        "",
    ),

    # ── NFM Land Use Changes (conditionally included) ─────────────────────────
    (
        "NFM Land Use Changes",
        "farmland_arable_before",
        "farmland_arable_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'enclosed_arable_farmland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='enclosed_arable_farmland'].areaBeforeHectares",
        "Hectares before land use change.",
    ),
    (
        "NFM Land Use Changes",
        "farmland_arable_after",
        "farmland_arable_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'enclosed_arable_farmland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='enclosed_arable_farmland'].areaAfterHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "farmland_livestock_before",
        "farmland_livestock_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'enclosed_livestock_farmland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='enclosed_livestock_farmland'].areaBeforeHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "farmland_livestock_after",
        "farmland_livestock_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'enclosed_livestock_farmland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='enclosed_livestock_farmland'].areaAfterHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "farmland_dairying_before",
        "farmland_dairying_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'enclosed_dairying_farmland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='enclosed_dairying_farmland'].areaBeforeHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "farmland_dairying_after",
        "farmland_dairying_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'enclosed_dairying_farmland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='enclosed_dairying_farmland'].areaAfterHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "semi_natural_grassland_before",
        "semi_natural_grassland_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'semi_natural_grassland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='semi_natural_grassland'].areaBeforeHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "semi_natural_grassland_after",
        "semi_natural_grassland_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'semi_natural_grassland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='semi_natural_grassland'].areaAfterHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "woodland_before",
        "woodland_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'woodland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='woodland'].areaBeforeHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "woodland_after",
        "woodland_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'woodland' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='woodland'].areaAfterHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "mountain_moors_before",
        "mountain_moors_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'mountain_moors_and_heath' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='mountain_moors_and_heath'].areaBeforeHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "mountain_moors_after",
        "mountain_moors_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'mountain_moors_and_heath' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='mountain_moors_and_heath'].areaAfterHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "peatland_restoration_before",
        "peatland_restoration_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'peatland_restoration' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='peatland_restoration'].areaBeforeHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "peatland_restoration_after",
        "peatland_restoration_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'peatland_restoration' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='peatland_restoration'].areaAfterHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "rivers_wetlands_before",
        "rivers_wetlands_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'rivers_wetlands_and_freshwater_habitats' must exist in "
        "pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='rivers_wetlands_and_freshwater_habitats']"
        ".areaBeforeHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "rivers_wetlands_after",
        "rivers_wetlands_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'rivers_wetlands_and_freshwater_habitats' must exist in "
        "pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='rivers_wetlands_and_freshwater_habitats']"
        ".areaAfterHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "coastal_margins_before",
        "coastal_margins_before",
        "decimal | null",
        "Optional",
        "Yes",
        "'coastal_margins' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='coastal_margins'].areaBeforeHectares",
        "",
    ),
    (
        "NFM Land Use Changes",
        "coastal_margins_after",
        "coastal_margins_after",
        "decimal | null",
        "Optional",
        "Yes",
        "'coastal_margins' must exist in pafs_core_nfm_land_use_changes",
        "Field OMITTED from payload",
        "nfmLandUseChanges[landUseType='coastal_margins'].areaAfterHectares",
        "",
    ),

    # ── Whole Life Costs ─────────────────────────────────────────────────────
    (
        "Whole Life Costs",
        "pv_appraisal_approach",
        "pv_appraisal_approach",
        "integer | null",
        "Optional",
        "Yes",
        "Submission validation requires non-null value only for "
        "MANDATORY_WL_TYPES (DEF, REF, REP). For HCR, ELO, STR, STU "
        "it is always sent but may be null.",
        "null",
        "project.wlcEstimatedWholeLifePvCosts",
        "Whole life present value of costs (£). Up to 18-digit integer.",
    ),
    (
        "Whole Life Costs",
        "pv_design_and_construction_costs",
        "pv_design_and_construction_costs",
        "integer | null",
        "Optional",
        "Yes",
        "Non-null required only for MANDATORY_WL_TYPES",
        "null",
        "project.wlcEstimatedDesignConstructionCosts",
        "PV design and construction costs (£).",
    ),
    (
        "Whole Life Costs",
        "pv_risk_contingency",
        "pv_risk_contingency",
        "integer | null",
        "Optional",
        "Yes",
        "Non-null required only for MANDATORY_WL_TYPES",
        "null",
        "project.wlcEstimatedRiskContingencyCosts",
        "PV risk contingency costs (£).",
    ),
    (
        "Whole Life Costs",
        "pv_future_costs",
        "pv_future_costs",
        "integer | null",
        "Optional",
        "Yes",
        "Non-null required only for MANDATORY_WL_TYPES",
        "null",
        "project.wlcEstimatedFutureCosts",
        "PV future maintenance costs (£).",
    ),
    (
        "Whole Life Costs",
        "pv_whole_life_benefits",
        "pv_whole_life_benefits",
        "integer | null",
        "Optional",
        "Yes",
        "Non-null required only for MANDATORY_WL_TYPES",
        "null",
        "project.wlbEstimatedWholeLifePvBenefits",
        "PV whole life benefits (£).",
    ),

    # ── Whole Life Benefits ──────────────────────────────────────────────────
    (
        "Whole Life Benefits",
        "property_damages_avoided",
        "property_damages_avoided",
        "integer | null",
        "Optional",
        "Yes",
        "Sent for all types; submission validation only enforces "
        "pv_whole_life_benefits for MANDATORY_WL_TYPES",
        "null",
        "project.wlbEstimatedPropertyDamagesAvoided",
        "Estimated property damages avoided (£).",
    ),
    (
        "Whole Life Benefits",
        "environmental_benefits",
        "environmental_benefits",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.wlbEstimatedEnvironmentalBenefits",
        "Estimated environmental benefits (£).",
    ),
    (
        "Whole Life Benefits",
        "recreation_and_tourism",
        "recreation_and_tourism",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.wlbEstimatedRecreationTourismBenefits",
        "Estimated recreation and tourism benefits (£).",
    ),
    (
        "Whole Life Benefits",
        "growth_and_regeneration_benefits",
        "growth_and_regeneration_benefits",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "project.wlbEstimatedLandValueUpliftBenefits",
        "Estimated land value uplift / regeneration benefits (£).",
    ),

    # ── Carbon Fields ────────────────────────────────────────────────────────
    (
        "Carbon",
        "capital_carbon",
        "capital_carbon",
        "decimal | null",
        "Required",
        "Yes",
        "CARBON_FREE_TYPES (STU, STR): value = computeCarbonResults("
        "project, fundingValues).constructionTotalFunding (calculated). "
        "All other types: value from project.carbonCostBuild.",
        "For STU/STR: constructionTotalFunding | null (calculated, not user-entered)",
        "project.carbonCostBuild OR computed constructionTotalFunding (STU/STR)",
        "Carbon cost of construction (tCO₂e). Up to 16 digits + 2 d.p.",
    ),
    (
        "Carbon",
        "carbon_operational_cost_forecast",
        "carbon_operational_cost_forecast",
        "integer | null",
        "Required",
        "Yes",
        "CARBON_FREE_TYPES (STU, STR): always sends null. "
        "All other types: from project field.",
        "null (for STU/STR)",
        "project.carbonOperationalCostForecast",
        "Operational carbon cost forecast. 18-digit integer.",
    ),
    (
        "Carbon",
        "carbon_lifecycle",
        "carbon_lifecycle",
        "decimal | null",
        "Optional",
        "Yes",
        "CARBON_FREE_TYPES (STU, STR): always sends 0 (not null). "
        "All other types: from project field.",
        "0 (for STU/STR)",
        "project.carbonCostOperation",
        "Lifecycle carbon (tCO₂e). Up to 16 digits + 2 d.p.",
    ),
    (
        "Carbon",
        "carbon_sequestered",
        "carbon_sequestered",
        "decimal | null",
        "Optional",
        "Yes",
        "CARBON_FREE_TYPES (STU, STR): always sends null.",
        "null (for STU/STR)",
        "project.carbonCostSequestered",
        "Carbon sequestered (tCO₂e).",
    ),
    (
        "Carbon",
        "carbon_avoided",
        "carbon_avoided",
        "decimal | null",
        "Optional",
        "Yes",
        "CARBON_FREE_TYPES (STU, STR): always sends null.",
        "null (for STU/STR)",
        "project.carbonCostAvoided",
        "Carbon avoided (tCO₂e).",
    ),
    (
        "Carbon",
        "carbon_net_economic_benefit",
        "carbon_net_economic_benefit",
        "integer | null",
        "Optional",
        "Yes",
        "CARBON_FREE_TYPES (STU, STR): always sends null.",
        "null (for STU/STR)",
        "project.carbonSavingsNetEconomicBenefit",
        "Net economic benefit of carbon savings (£). 18-digit integer.",
    ),

    # ── Funding Sources ──────────────────────────────────────────────────────
    (
        "Funding Sources",
        "funding_sources",
        "funding_sources",
        "object",
        "Required",
        "No",
        "—",
        "{ values: [] }",
        "pafs_core_funding_values + pafs_core_funding_contributors",
        "Wrapper object containing the 'values' array of funding year entries.",
    ),
    (
        "Funding Sources — Year Entry",
        "financial_year",
        "funding_sources.values[n].financial_year",
        "integer",
        "Required",
        "No",
        "—",
        "n/a (entry not created)",
        "pafs_core_funding_values[].financialYear (parsed from 'YYYY/YY')",
        "The start year of the financial year (e.g. 2025 for 2025/26).",
    ),
    (
        "Funding Sources — Year Entry",
        "fcerm_gia",
        "funding_sources.values[n].fcerm_gia",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].fcermGia",
        "FCERM Grant in Aid (£).",
    ),
    (
        "Funding Sources — Year Entry",
        "asset_replacement_allowance",
        "funding_sources.values[n].asset_replacement_allowance",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].assetReplacementAllowance",
        "ARA funding (£).",
    ),
    (
        "Funding Sources — Year Entry",
        "environment_statutory_funding",
        "funding_sources.values[n].environment_statutory_funding",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].environmentStatutoryFunding",
        "Environment statutory funding (£).",
    ),
    (
        "Funding Sources — Year Entry",
        "frequently_floodded_communities",
        "funding_sources.values[n].frequently_floodded_communities",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].frequentlyFloodedCommunities",
        "Note: typo 'floodded' is intentional in the payload field name.",
    ),
    (
        "Funding Sources — Year Entry",
        "other_additional_grant_in_aid",
        "funding_sources.values[n].other_additional_grant_in_aid",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].otherAdditionalGrantInAid",
        "",
    ),
    (
        "Funding Sources — Year Entry",
        "other_government_department",
        "funding_sources.values[n].other_government_department",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].otherGovernmentDepartment",
        "",
    ),
    (
        "Funding Sources — Year Entry",
        "recovery",
        "funding_sources.values[n].recovery",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].recovery",
        "",
    ),
    (
        "Funding Sources — Year Entry",
        "summer_economic_fund",
        "funding_sources.values[n].summer_economic_fund",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].summerEconomicFund",
        "",
    ),
    (
        "Funding Sources — Year Entry",
        "local_levy",
        "funding_sources.values[n].local_levy",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].localLevy",
        "",
    ),
    (
        "Funding Sources — Year Entry",
        "internal_drainage_boards",
        "funding_sources.values[n].internal_drainage_boards",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].internalDrainageBoards",
        "",
    ),
    (
        "Funding Sources — Year Entry",
        "not_yet_identified",
        "funding_sources.values[n].not_yet_identified",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_values[].notYetIdentified",
        "",
    ),
    (
        "Funding Sources — Contributors",
        "public_contributions",
        "funding_sources.values[n].public_contributions",
        "array",
        "Optional",
        "No",
        "—",
        "[] (empty array)",
        "pafs_core_funding_contributors filtered by contributorType='public_contributions'",
        "Array of { name, amount } contributor entries.",
    ),
    (
        "Funding Sources — Contributors",
        "private_contributions",
        "funding_sources.values[n].private_contributions",
        "array",
        "Optional",
        "No",
        "—",
        "[] (empty array)",
        "pafs_core_funding_contributors filtered by contributorType='private_contributions'",
        "",
    ),
    (
        "Funding Sources — Contributors",
        "other_ea_contributions",
        "funding_sources.values[n].other_ea_contributions",
        "array",
        "Optional",
        "No",
        "—",
        "[] (empty array)",
        "pafs_core_funding_contributors filtered by contributorType='other_ea_contributions'",
        "",
    ),
    (
        "Funding Sources — Contributor Entry",
        "name",
        "funding_sources.values[n].<contributor_type>[m].name",
        "string | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_contributors[].name",
        "Name of the contributing organisation.",
    ),
    (
        "Funding Sources — Contributor Entry",
        "amount",
        "funding_sources.values[n].<contributor_type>[m].amount",
        "integer | null",
        "Optional",
        "No",
        "—",
        "null",
        "pafs_core_funding_contributors[].amount",
        "Contribution amount (£). 18-digit integer.",
    ),
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark navy
SECTION_FILLS = {
    "Core Identity":                   PatternFill("solid", fgColor="DEEAF1"),
    "Gateway Dates":                   PatternFill("solid", fgColor="E2EFDA"),
    "Secondary Risk Sources":          PatternFill("solid", fgColor="FFF2CC"),
    "Risk & Properties":               PatternFill("solid", fgColor="FCE4D6"),
    "Goals & Approach":                PatternFill("solid", fgColor="EAD1DC"),
    "Outcome Measures — OM2":          PatternFill("solid", fgColor="D9E1F2"),
    "Outcome Measures — OM3":          PatternFill("solid", fgColor="D9E1F2"),
    "Outcome Measures — OM4a":         PatternFill("solid", fgColor="D9E1F2"),
    "Outcome Measures — OM4b":         PatternFill("solid", fgColor="D9E1F2"),
    "Confidence":                      PatternFill("solid", fgColor="E2EFDA"),
    "NFM Details":                     PatternFill("solid", fgColor="FFF2CC"),
    "NFM Measures":                    PatternFill("solid", fgColor="FFF2CC"),
    "NFM Land Use Changes":            PatternFill("solid", fgColor="FFF2CC"),
    "Whole Life Costs":                PatternFill("solid", fgColor="FCE4D6"),
    "Whole Life Benefits":             PatternFill("solid", fgColor="FCE4D6"),
    "Carbon":                          PatternFill("solid", fgColor="EAD1DC"),
    "Funding Sources":                 PatternFill("solid", fgColor="DEEAF1"),
    "Funding Sources — Year Entry":    PatternFill("solid", fgColor="DEEAF1"),
    "Funding Sources — Contributors":  PatternFill("solid", fgColor="DEEAF1"),
    "Funding Sources — Contributor Entry": PatternFill("solid", fgColor="DEEAF1"),
}
DEFAULT_FILL  = PatternFill("solid", fgColor="FFFFFF")

YES_FILL      = PatternFill("solid", fgColor="FFF2CC")   # amber for Conditional=Yes
REQ_FILL      = PatternFill("solid", fgColor="E2EFDA")   # light green for Required

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def _header_font():
    return Font(name="Calibri", bold=True, color="FFFFFF", size=11)

def _body_font(bold=False):
    return Font(name="Calibri", bold=bold, size=10)

def _apply_border(cell):
    cell.border = BORDER

def _wrap(cell):
    cell.alignment = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

def build_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Submission Payload Analysis"

    # ── Headers ──────────────────────────────────────────────────────────────
    ws.append(COLUMNS)
    header_row = ws[1]
    for cell in header_row:
        cell.fill = HEADER_FILL
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        _apply_border(cell)
    ws.row_dimensions[1].height = 36

    # ── Data rows ─────────────────────────────────────────────────────────────
    prev_section = None
    for row_idx, row_data in enumerate(ROWS, start=2):
        ws.append(list(row_data))
        section = row_data[0]
        conditional = row_data[5]
        schema_req   = row_data[4]

        row_fill = SECTION_FILLS.get(section, DEFAULT_FILL)

        for col_idx, cell in enumerate(ws[row_idx], start=1):
            _apply_border(cell)
            _wrap(cell)
            cell.font = _body_font()

            # Default: light section colour
            cell.fill = row_fill

            # Highlight Conditional column
            if col_idx == 6:   # F — Conditional?
                if conditional == "Yes":
                    cell.fill = YES_FILL
                    cell.font = _body_font(bold=True)

            # Highlight Required column
            if col_idx == 5:   # E — Schema Required?
                if "Required" in schema_req and "Optional" not in schema_req:
                    cell.fill = REQ_FILL

        ws.row_dimensions[row_idx].height = 48

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [28, 40, 52, 18, 22, 14, 55, 40, 60, 60]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Freeze pane ───────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    return wb


# ---------------------------------------------------------------------------
# Legend sheet
# ---------------------------------------------------------------------------

def add_legend(wb):
    ls = wb.create_sheet("Legend")
    legend_data = [
        ("Column", "Description"),
        ("A — Section", "Logical grouping of payload fields"),
        ("B — Property Name", "The exact JSON key as sent in the payload"),
        ("C — Full Payload Path", "Dot-notation path from the root payload object"),
        ("D — Data Type",
         "string / integer / decimal / boolean / array / object / null variants"),
        ("E — Schema Required?",
         "Whether Joi schema marks the field Required or Optional. "
         "'Required (nullable)' means the key must be present but the value may be null."),
        ("F — Conditional?",
         "Yes = the VALUE sent (or whether the field is included at all) "
         "depends on a runtime condition. No = always sent with the same logic."),
        ("G — Condition Description",
         "The runtime condition that determines what value is sent"),
        ("H — Value When Condition Not Met",
         "'null' = field is sent with JSON null. "
         "'false' = field is sent as boolean false. "
         "'0' = field is sent as zero. "
         "'Field OMITTED' = the key is not included in the payload at all. "
         "'n/a' = condition cannot be not-met (always present)."),
        ("I — Source Field(s) in Project Object",
         "The project object property (or computed value) the payload field comes from"),
        ("J — Notes / Allowed Values",
         "Additional context, allowed enum values, label maps, units, etc."),
        ("", ""),
        ("Project Type Abbreviations", ""),
        ("DEF", "Defence — full MANDATORY_WL_TYPES checklist required"),
        ("REF", "Refurbishment — full MANDATORY_WL_TYPES checklist required"),
        ("REP", "Replacement — full MANDATORY_WL_TYPES checklist required"),
        ("HCR", "Habitat Creation/Restoration — OPTIONAL_WL_TYPES"),
        ("ELO", "Environmental Land Objective — OPTIONAL_WL_TYPES"),
        ("STR", "Strategy — CARBON_FREE_TYPES (no carbon user input)"),
        ("STU", "Study — CARBON_FREE_TYPES (no carbon user input)"),
        ("", ""),
        ("MANDATORY_WL_TYPES", "DEF, REF, REP — require WLC, WLB, confidence and NFM sections"),
        ("OPTIONAL_WL_TYPES",  "HCR, ELO — WLC/WLB optional"),
        ("CARBON_FREE_TYPES",  "STU, STR — capital_carbon calculated; no user carbon fields"),
    ]

    for r_idx, (col_a, col_b) in enumerate(legend_data, start=1):
        ls.cell(r_idx, 1).value = col_a
        ls.cell(r_idx, 2).value = col_b

        if r_idx == 1:
            for c in [ls.cell(r_idx, 1), ls.cell(r_idx, 2)]:
                c.fill = HEADER_FILL
                c.font = _header_font()
                c.alignment = Alignment(vertical="top", wrap_text=True)
        else:
            ls.cell(r_idx, 1).font = _body_font(bold=True)
            ls.cell(r_idx, 2).font = _body_font()
            ls.cell(r_idx, 2).alignment = Alignment(wrap_text=True, vertical="top")

        ls.row_dimensions[r_idx].height = 30

    ls.column_dimensions["A"].width = 45
    ls.column_dimensions["B"].width = 80


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

OUTPUT_PATH = (
    "/Users/795110/Documents/Fujale-Space/Work/Documents/FCRM/PAFs/"
    "Additional Documents/pafs_submission_payload_analysis.xlsx"
)

if __name__ == "__main__":
    wb = build_workbook()
    add_legend(wb)
    wb.save(OUTPUT_PATH)
    print(f"Saved: {OUTPUT_PATH}")
    print(f"Total data rows: {len(ROWS)}")
